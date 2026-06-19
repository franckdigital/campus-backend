from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from django.http import HttpResponse
from django.db.models import Sum, Count
from django.db import models
import openpyxl
from io import BytesIO

from .models import AccountingAccount, JournalEntry, JournalLine
from .serializers import (
    AccountingAccountSerializer, JournalEntrySerializer,
    JournalEntryListSerializer, JournalEntryCreateSerializer,
    JournalLineSerializer
)


OHADA_ACCOUNTS = [
    # Trésorerie & banque (ASSET)
    {'code': '571', 'name': 'Caisse', 'account_type': 'ASSET'},
    {'code': '521', 'name': 'Banque', 'account_type': 'ASSET'},
    # Créances clients (ASSET)
    {'code': '411', 'name': 'Étudiants – créances scolarité', 'account_type': 'ASSET'},
    {'code': '412', 'name': 'Clients douteux', 'account_type': 'ASSET'},
    # Passif
    {'code': '401', 'name': 'Fournisseurs', 'account_type': 'LIABILITY'},
    {'code': '421', 'name': 'Personnel – rémunérations dues', 'account_type': 'LIABILITY'},
    # Capitaux propres
    {'code': '101', 'name': 'Capital social', 'account_type': 'EQUITY'},
    {'code': '111', 'name': 'Réserves légales', 'account_type': 'EQUITY'},
    # Produits (REVENUE)
    {'code': '706', 'name': 'Frais de scolarité', 'account_type': 'REVENUE'},
    {'code': '7061', 'name': "Frais d'inscription", 'account_type': 'REVENUE'},
    {'code': '708', 'name': 'Autres produits d\'activité', 'account_type': 'REVENUE'},
    # Charges (EXPENSE)
    {'code': '621', 'name': 'Rémunérations – personnel enseignant', 'account_type': 'EXPENSE'},
    {'code': '622', 'name': 'Rémunérations – personnel administratif', 'account_type': 'EXPENSE'},
    {'code': '631', 'name': 'Loyers et charges locatives', 'account_type': 'EXPENSE'},
    {'code': '635', 'name': 'Entretien et réparations', 'account_type': 'EXPENSE'},
    {'code': '641', 'name': 'Fournitures scolaires et bureautiques', 'account_type': 'EXPENSE'},
    {'code': '651', 'name': 'Marketing et publicité', 'account_type': 'EXPENSE'},
    {'code': '671', 'name': 'Charges exceptionnelles', 'account_type': 'EXPENSE'},
]


class AccountingAccountViewSet(viewsets.ModelViewSet):
    queryset = AccountingAccount.objects.select_related('site', 'parent').all()
    serializer_class = AccountingAccountSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['code', 'name']
    ordering_fields = ['code', 'name']
    filterset_fields = ['site', 'account_type', 'parent', 'is_active']


class JournalEntryViewSet(viewsets.ModelViewSet):
    queryset = JournalEntry.objects.select_related(
        'site', 'payment', 'created_by', 'posted_by'
    ).prefetch_related('lines__account').all()
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['entry_number', 'description', 'reference']
    ordering_fields = ['entry_date', 'entry_number']
    filterset_fields = ['site', 'status', 'entry_date', 'is_active']

    def get_serializer_class(self):
        if self.action == 'list':
            return JournalEntryListSerializer
        if self.action == 'create':
            return JournalEntryCreateSerializer
        return JournalEntrySerializer

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=['post'])
    def post(self, request, pk=None):
        journal_entry = self.get_object()
        if journal_entry.post(request.user):
            return Response(JournalEntrySerializer(journal_entry).data)
        return Response(
            {'detail': 'Impossible de valider cette écriture (non équilibrée ou déjà validée)'},
            status=status.HTTP_400_BAD_REQUEST
        )

    @action(detail=True, methods=['post'], url_path='add-line')
    def add_line(self, request, pk=None):
        journal_entry = self.get_object()
        if journal_entry.status != 'DRAFT':
            return Response(
                {'detail': 'Impossible de modifier une écriture validée'},
                status=status.HTTP_400_BAD_REQUEST
            )
        serializer = JournalLineSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(journal_entry=journal_entry)
        return Response(JournalEntrySerializer(journal_entry).data)


class JournalLineViewSet(viewsets.ModelViewSet):
    queryset = JournalLine.objects.select_related('journal_entry', 'account').all()
    serializer_class = JournalLineSerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['journal_entry', 'account', 'is_active']


class AccountingExportView(APIView):
    def get(self, request):
        site_id = request.query_params.get('site_id')
        period = request.query_params.get('period')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        entries = JournalEntry.objects.filter(status='POSTED')

        if site_id:
            entries = entries.filter(site_id=site_id)

        if period:
            from datetime import datetime
            try:
                year, month = period.split('-')
                entries = entries.filter(entry_date__year=int(year), entry_date__month=int(month))
            except ValueError:
                pass
        elif start_date and end_date:
            entries = entries.filter(entry_date__range=(start_date, end_date))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Journal Comptable"

        # Style header
        headers = ['N° Écriture', 'Date', 'Description', 'Référence', 'Compte', 'Libellé ligne', 'Débit', 'Crédit']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)

        row = 2
        for entry in entries.prefetch_related('lines__account'):
            for line in entry.lines.filter(is_active=True):
                ws.cell(row=row, column=1, value=entry.entry_number)
                ws.cell(row=row, column=2, value=entry.entry_date.strftime('%d/%m/%Y'))
                ws.cell(row=row, column=3, value=entry.description)
                ws.cell(row=row, column=4, value=entry.reference)
                ws.cell(row=row, column=5, value=line.account.code)
                ws.cell(row=row, column=6, value=line.description or line.account.name)
                ws.cell(row=row, column=7, value=float(line.debit_amount))
                ws.cell(row=row, column=8, value=float(line.credit_amount))
                row += 1

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        label = period or (f"{start_date}_{end_date}" if start_date else 'all')
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="journal_comptable_{label}.xlsx"'
        return response


class TrialBalanceView(APIView):
    def get(self, request):
        site_id = request.query_params.get('site_id')

        accounts = AccountingAccount.objects.filter(is_active=True)
        if site_id:
            accounts = accounts.filter(site_id=site_id)

        balances = []
        for account in accounts.order_by('code'):
            # Include DRAFT entries so auto-generated F1 journal entries show immediately
            debit_total = account.debit_lines.filter(
                is_active=True, journal_entry__status__in=['DRAFT', 'POSTED']
            ).aggregate(total=Sum('debit_amount'))['total'] or 0

            credit_total = account.credit_lines.filter(
                is_active=True, journal_entry__status__in=['DRAFT', 'POSTED']
            ).aggregate(total=Sum('credit_amount'))['total'] or 0

            if debit_total > 0 or credit_total > 0:
                balances.append({
                    'account_code': account.code,
                    'account_name': account.name,
                    'account_type': account.account_type,
                    'debit_total': float(debit_total),
                    'credit_total': float(credit_total),
                    'balance': float(debit_total - credit_total),
                })

        total_debit = sum(b['debit_total'] for b in balances)
        total_credit = sum(b['credit_total'] for b in balances)

        return Response({
            'accounts': balances,
            'total_debit': total_debit,
            'total_credit': total_credit,
            'is_balanced': abs(total_debit - total_credit) < 0.01,
        })


class RevenueReportView(APIView):
    """F4: Revenue by period and by class/filière — based on Invoice.amount_paid."""
    def get(self, request):
        from apps.finance.models import Invoice
        from django.db.models.functions import TruncMonth
        from django.utils import timezone
        from datetime import date

        site_id = request.query_params.get('site_id')
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        report_type = request.query_params.get('type', 'monthly')

        now = timezone.now().date()
        try:
            start_date = date.fromisoformat(start_date_str) if start_date_str else now.replace(month=1, day=1)
        except ValueError:
            start_date = now.replace(month=1, day=1)
        try:
            end_date = date.fromisoformat(end_date_str) if end_date_str else now
        except ValueError:
            end_date = now

        # Use Invoice.amount_paid — always accurate regardless of payment status workflow
        invoices_qs = Invoice.objects.filter(
            amount_paid__gt=0,
            is_active=True,
            issue_date__range=(start_date, end_date),
        )
        if site_id:
            invoices_qs = invoices_qs.filter(site_id=site_id)

        total = float(invoices_qs.aggregate(t=Sum('amount_paid'))['t'] or 0)

        if report_type == 'monthly':
            data = (
                invoices_qs
                .annotate(month=TruncMonth('issue_date'))
                .values('month')
                .annotate(total=Sum('amount_paid'))
                .order_by('month')
            )
            return Response({
                'type': 'monthly',
                'total': total,
                'data': [
                    {'month': r['month'].strftime('%b %Y'), 'total': float(r['total'])}
                    for r in data
                ],
            })

        elif report_type == 'by_class':
            from apps.academic.models import Enrollment

            student_ids = list(invoices_qs.values_list('student_id', flat=True).distinct())
            enrollments = Enrollment.objects.filter(
                student_id__in=student_ids, is_active=True
            ).select_related('class_obj')
            student_class_map = {str(e.student_id): e.class_obj.name for e in enrollments}

            class_totals = {}
            for inv in invoices_qs.select_related('student'):
                sid = str(inv.student_id)
                cls = student_class_map.get(sid, 'Non classé')
                class_totals[cls] = class_totals.get(cls, 0) + float(inv.amount_paid)

            data = sorted(
                [{'class_name': k, 'total': v} for k, v in class_totals.items()],
                key=lambda x: -x['total'],
            )
            return Response({'type': 'by_class', 'total': total, 'data': data})

        return Response({'detail': 'Type invalide'}, status=status.HTTP_400_BAD_REQUEST)


class UnpaidReportView(APIView):
    """F4: Unpaid invoices report."""
    def get(self, request):
        from apps.finance.models import Invoice
        from django.utils import timezone

        site_id = request.query_params.get('site_id')

        unpaid = Invoice.objects.filter(
            status__in=['SENT', 'PARTIAL', 'OVERDUE'],
            is_active=True
        ).select_related('student__user', 'academic_year')

        if site_id:
            unpaid = unpaid.filter(site_id=site_id)

        now = timezone.now().date()
        agg = unpaid.aggregate(total_balance=Sum('balance'), cnt=Count('id'))
        overdue_count = unpaid.filter(due_date__lt=now).count()

        data = [
            {
                'id': str(inv.id),
                'invoice_number': inv.invoice_number,
                'student_name': inv.student.user.full_name,
                'student_matricule': inv.student.matricule,
                'total': float(inv.total),
                'amount_paid': float(inv.amount_paid),
                'balance': float(inv.balance),
                'due_date': str(inv.due_date) if inv.due_date else None,
                'status': inv.status,
                'is_overdue': bool(inv.due_date and inv.due_date < now),
            }
            for inv in unpaid.order_by('-balance')
        ]

        return Response({
            'invoices': data,
            'total_balance': float(agg['total_balance'] or 0),
            'count': agg['cnt'],
            'overdue_count': overdue_count,
        })


class InitOHADAView(APIView):
    """F2: Seed default OHADA chart of accounts for a site."""
    def post(self, request):
        from apps.core.models import Site

        site_id = request.data.get('site_id')
        if site_id:
            try:
                site = Site.objects.get(id=site_id)
            except Site.DoesNotExist:
                return Response({'detail': 'Site non trouvé'}, status=status.HTTP_404_NOT_FOUND)
        else:
            site = Site.objects.filter(is_active=True).first()

        if not site:
            return Response({'detail': 'Aucun site disponible'}, status=status.HTTP_400_BAD_REQUEST)

        created = 0
        for acc in OHADA_ACCOUNTS:
            _, was_created = AccountingAccount.objects.get_or_create(
                code=acc['code'], site=site,
                defaults={**acc, 'is_system': True, 'is_active': True, 'description': ''}
            )
            if was_created:
                created += 1

        return Response({
            'detail': f'{created} compte(s) OHADA initialisé(s) pour {site.name}',
            'created': created,
            'site_name': site.name,
        })
